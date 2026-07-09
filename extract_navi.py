# -*- coding: utf-8 -*-
"""
경기도교육청 수시NAVI(수시나비) 엑셀 → navi.json 추출기.

원칙(생기부분석기와 동일): 실측·참고 정보만. 합격률/등급컷/지원데이터는 의도적으로 제외.
추출 대상:
  - '종합' 시트: 학종 전형별 평가요소 역량비중(학업/진로/공동체) + 인재상 + 면접 + 수능최저 (공개 모집요강 수준)
  - '교과반영' 시트: 진로선택과목 반영 '경향'만 집계(원본 행 임베드 안 함)

사용: python extract_navi.py [원본.xlsx] [navi.json]
원본이 암호화(OOXML, 기본암호 VelvetSweatshop)면 자동 복호화.
"""
import sys, io, json, datetime
import openpyxl

DEFAULT_SRC = r"C:/Users/user/Desktop/경기도교육청_2027수시NAVI(수시나비)_260707.xlsx"
DEFAULT_OUT = r"C:/Users/user/claude/sgb/navi.json"


def load_wb(path):
    """암호화 xlsx면 복호화 후 로드."""
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        import msoffcrypto
        buf = io.BytesIO()
        with open(path, "rb") as f:
            off = msoffcrypto.OfficeFile(f)
            off.load_key(password="VelvetSweatshop")  # 엑셀 기본암호 = 사실상 무암호
            off.decrypt(buf)
        buf.seek(0)
        return openpyxl.load_workbook(buf, read_only=True, data_only=True)


def num(v):
    return v if isinstance(v, (int, float)) else None


def s(v):
    if v is None:
        return None
    t = str(v).replace("\n", " ").strip()
    return t or None


def extract_jonghap(ws):
    """'종합' 시트 → 전형 레코드 목록. 병합셀(지역/대학) forward-fill."""
    prev = {}

    def ff(v, key):
        t = s(v)
        if t and t != "*":
            prev[key] = t
            return t
        return prev.get(key)

    out = []
    for r in ws.iter_rows(min_row=8, values_only=True):
        지역 = ff(r[1], "region")
        대학 = ff(r[2], "univ")
        전형 = s(r[3])
        if not 전형 or 전형 == "*":
            continue
        rec = {
            "region": 지역, "univ": 대학, "type": 전형, "track": s(r[4]),
            "quota": num(r[5]), "stage": s(r[6]), "method": s(r[8]),
            "interview": s(r[12]), "minGrade": s(r[13]), "minDetail": s(r[14]),
            "acad": num(r[15]), "career": num(r[16]), "community": num(r[17]),
            "evalNote": s(r[18]), "ideal": s(r[19]),
        }
        out.append(rec)
    return out


def extract_gyogwa_trend(ws):
    """'교과반영' 시트 → 진로선택과목 반영 '경향'만 집계(원본 미임베드)."""
    univ_reflect = set()
    univ_all = set()
    for r in ws.iter_rows(min_row=7, values_only=True):
        univ = s(r[3])
        if not univ or univ == "*":
            continue
        univ_all.add(univ)
        # 진로선택 반영교과(col15) 비어있지 않고 '미반영' 아니면 반영으로 집계
        jinro_area = s(r[15])
        if jinro_area and jinro_area != "미반영":
            univ_reflect.add(univ)
    return {
        "univTotal": len(univ_all),
        "univReflectJinro": len(univ_reflect),
    }


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SRC
    out = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUT
    wb = load_wb(src)
    jonghap = extract_jonghap(wb["종합"])
    gyogwa = extract_gyogwa_trend(wb["교과반영"])
    data = {
        "meta": {
            "source": "경기도교육청 2027 수시NAVI(수시나비)",
            "extracted": datetime.date.today().isoformat(),
            "note": "학종 평가요소 역량비중·인재상·면접·수능최저 등 공개 전형정보만 추출. "
                    "합격률·등급컷·지원데이터는 원칙(합격예측 거절)에 따라 제외.",
        },
        "jonghap": jonghap,
        "gyogwa": gyogwa,
    }
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    numw = sum(1 for x in jonghap if isinstance(x["acad"], (int, float)))
    print(f"navi.json 생성: 전형 {len(jonghap)}건 "
          f"(역량비중 숫자 {numw}건, 대학 {len({x['univ'] for x in jonghap if x['univ']})}곳) "
          f"| 교과 진로선택 반영 {gyogwa['univReflectJinro']}/{gyogwa['univTotal']}곳")


if __name__ == "__main__":
    main()
